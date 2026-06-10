import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone

def style_header_cell(cell, font_size=11, bold=True):
    cell.font = Font(name='Segoe UI', size=font_size, bold=bold, color='FFFFFF')
    cell.fill = PatternFill(start_color='003388', end_color='003388', fill_type='solid') # Deep Blue ishelp.co.uk
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_sub_header_cell(cell, font_size=11, bold=True):
    cell.font = Font(name='Segoe UI', size=font_size, bold=bold, color='FFFFFF')
    cell.fill = PatternFill(start_color='882211', end_color='882211', fill_type='solid') # Rust Red ishelp.co.uk
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_data_cell(cell, horizontal='left', bold=False, italic=False):
    cell.font = Font(name='Segoe UI', size=10, bold=bold, italic=italic)
    cell.alignment = Alignment(horizontal=horizontal, vertical='center')

def set_gridlines_and_auto_width(ws):
    ws.views.sheetView[0].showGridLines = True
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if '\n' in val_str:
                lines = val_str.split('\n')
                max_len = max(max_len, max(len(l) for l in lines))
            else:
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def generate_xlsx_report(assessment, report_type):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    if report_type == 'ExecutiveSummary':
        ws.title = "Executive Summary"
        
        # Title Block
        ws.merge_cells("A1:D1")
        ws["A1"] = f"Executive Summary: {assessment.name}"
        ws["A1"].font = Font(name='Segoe UI', size=16, bold=True, color='003388')
        ws.row_dimensions[1].height = 30
        
        # Meta details
        ws["A3"] = "Client Name:"
        ws["B3"] = assessment.client.name
        ws["A4"] = "Client Email:"
        ws["B4"] = assessment.client.email
        ws["A5"] = "Assessment Date:"
        ws["B5"] = timezone.now().strftime("%Y-%m-%d")
        ws["A6"] = "Methodology Version:"
        ws["B6"] = str(assessment.methodology_version)
        ws["A7"] = "Assessment Status:"
        ws["B7"] = assessment.status

        for row in range(3, 8):
            ws[f"A{row}"].font = Font(name='Segoe UI', size=10, bold=True)
            ws[f"B{row}"].font = Font(name='Segoe UI', size=10)

        # Risk Metrics Table
        ws["A9"] = "Risk Posture Metrics"
        ws.merge_cells("A9:B9")
        ws["A9"].font = Font(name='Segoe UI', size=12, bold=True, color='882211')

        headers = ["Metric", "Value"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=10, column=col_idx, value=h)
            style_header_cell(cell)

        metrics = [
            ("Total Registered Risks", assessment.risk_items.count()),
            ("High/Critical Risks", len([r for r in assessment.risk_items.all() if r.risk_category in ['High', 'Critical']])),
            ("Open Gaps/Findings", assessment.findings.filter(status='Open').count()),
            ("Mitigated Risks", assessment.risk_items.filter(treatment__status__in=['Mitigated', 'Closed']).count()),
        ]

        for idx, (m, v) in enumerate(metrics, start=11):
            cell_m = ws.cell(row=idx, column=1, value=m)
            cell_v = ws.cell(row=idx, column=2, value=v)
            style_data_cell(cell_m)
            style_data_cell(cell_v, horizontal='right')
            cell_m.border = thin_border
            cell_v.border = thin_border

    elif report_type == 'DetailedRiskAssessment':
        ws.title = "Detailed Assessment"
        
        ws.merge_cells("A1:H1")
        ws["A1"] = f"Detailed Cyber Risk Assessment: {assessment.name}"
        ws["A1"].font = Font(name='Segoe UI', size=16, bold=True, color='003388')
        ws.row_dimensions[1].height = 30

        # General details
        details = [
            ("Client Name", assessment.client.name),
            ("Client Contact", assessment.client.email),
            ("Methodology Version", str(assessment.methodology_version)),
            ("Assessment Status", assessment.status),
            ("Main Asset Evaluated", assessment.asset),
            ("Asset Location", assessment.location),
            ("Asset Owner", assessment.owner),
        ]
        for idx, (k, v) in enumerate(details, start=3):
            cell_k = ws.cell(row=idx, column=1, value=k)
            cell_v = ws.cell(row=idx, column=2, value=v)
            style_data_cell(cell_k, bold=True)
            style_data_cell(cell_v)

        # Risk items table
        start_row = 12
        ws.cell(row=start_row - 1, column=1, value="Identified Risk Items").font = Font(name='Segoe UI', size=12, bold=True, color='882211')
        
        headers = [
            "Asset", "Location", "Owner", "Threat Scenario", "Vulnerability details",
            "Inherent Score", "Inherent Category", "Treatment Action", "Residual Score", "Residual Category"
        ]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=h)
            style_header_cell(cell)

        row_idx = start_row + 1
        for item in assessment.risk_items.all():
            treatment_action = getattr(item, 'treatment', None)
            action_text = treatment_action.action if treatment_action else "No treatment logged"
            
            data = [
                item.asset_name, item.asset_location, item.asset_owner,
                item.threat.name, item.vulnerability,
                item.risk_score, item.risk_category,
                action_text,
                item.residual_risk_score if item.residual_risk_score is not None else "N/A",
                item.residual_risk_category
            ]
            for col_idx, val in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                style_data_cell(cell)
                cell.border = thin_border
            row_idx += 1

    elif report_type == 'RiskRegister':
        ws.title = "Risk Register"
        
        ws.merge_cells("A1:K1")
        ws["A1"] = f"Corporate Cyber Risk Register - {assessment.name}"
        ws["A1"].font = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
        ws["A1"].fill = PatternFill(start_color='003388', end_color='003388', fill_type='solid')
        ws["A1"].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        headers = [
            "ID", "Asset Affected", "Asset Location", "Asset Owner", 
            "Threat Vector", "Vulnerability Details", "Existing Controls",
            "Inherent Score", "Inherent Category", "Residual Score", "Residual Category"
        ]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=h)
            style_header_cell(cell)

        row_idx = 4
        for item in assessment.risk_items.all():
            data = [
                item.id, item.asset_name, item.asset_location, item.asset_owner,
                item.threat.name, item.vulnerability, item.existing_controls,
                item.risk_score, item.risk_category,
                item.residual_risk_score if item.residual_risk_score is not None else "N/A",
                item.residual_risk_category
            ]
            for col_idx, val in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                style_data_cell(cell)
                cell.border = thin_border
            row_idx += 1

    elif report_type == 'TreatmentPlan':
        ws.title = "Treatment Plan"
        
        ws.merge_cells("A1:F1")
        ws["A1"] = f"Risk Treatment Plan: {assessment.name}"
        ws["A1"].font = Font(name='Segoe UI', size=16, bold=True, color='003388')
        ws.row_dimensions[1].height = 30

        headers = ["Asset Affected", "Threat Vector", "Treatment Action", "Owner", "Target Date", "Status"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=h)
            style_header_cell(cell)

        row_idx = 4
        for item in assessment.risk_items.all():
            treatment = getattr(item, 'treatment', None)
            if treatment:
                data = [
                    item.asset_name,
                    item.threat.name,
                    treatment.action,
                    treatment.owner,
                    treatment.target_date.strftime("%Y-%m-%d") if treatment.target_date else "N/A",
                    treatment.status
                ]
                for col_idx, val in enumerate(data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    style_data_cell(cell)
                    cell.border = thin_border
                row_idx += 1

    elif report_type == 'ChangeRequest':
        ws.title = "Change Request Review"
        
        ws.merge_cells("A1:D1")
        ws["A1"] = f"Change Request Assessment: {assessment.name}"
        ws["A1"].font = Font(name='Segoe UI', size=16, bold=True, color='003388')
        ws.row_dimensions[1].height = 30

        ws["A3"] = "Change Request Reference / Context:"
        ws["A3"].font = Font(name='Segoe UI', size=11, bold=True)
        ws.merge_cells("A4:D6")
        ws["A4"] = assessment.change_request or "No associated change request details provided."
        ws["A4"].alignment = Alignment(vertical='top', wrap_text=True)

        ws["A8"] = "Business Process Impact:"
        ws["A8"].font = Font(name='Segoe UI', size=11, bold=True)
        ws.merge_cells("A9:D11")
        ws["A9"] = assessment.business_process_impact or "No impact assessment provided."
        ws["A9"].alignment = Alignment(vertical='top', wrap_text=True)

        ws["A13"] = "CIA Security Impact Assessment:"
        ws["A13"].font = Font(name='Segoe UI', size=11, bold=True)
        ws["A14"] = "Confidentiality Affected:"
        ws["B14"] = "Yes" if assessment.confidentiality_affected else "No"
        ws["A15"] = "Integrity Affected:"
        ws["B15"] = "Yes" if assessment.integrity_affected else "No"
        ws["A16"] = "Availability Affected:"
        ws["B16"] = "Yes" if assessment.availability_affected else "No"

        for r in range(14, 17):
            ws[f"A{r}"].font = Font(name='Segoe UI', size=10, bold=True)
            ws[f"B{r}"].font = Font(name='Segoe UI', size=10)

    set_gridlines_and_auto_width(ws)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
